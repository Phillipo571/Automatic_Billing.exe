import sys, os, threading, time, gc, uuid, shutil, urllib.parse, pathlib, tempfile
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import pythoncom
import win32com.client

from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QFileDialog,
    QMessageBox, QProgressBar, QGridLayout, QHBoxLayout, QDialog
)
from PySide6.QtCore import Qt, Signal, QObject
from PySide6.QtGui import QIcon

class WorkerSignals(QObject):
    progress = Signal(int)
    finished = Signal(str)
    error = Signal(str)

class BillingWorker(threading.Thread):
    def __init__(self, customer, file_path, signals):
        super().__init__()
        self.customer = customer
        self.file_path = file_path
        self.signals = signals
        self.stop_requested = False
        self.cancel_mode = False
        self.temp_output_file = os.path.join(tempfile.gettempdir(), f"{customer}_temp_result.xlsx")

        if os.path.exists(self.temp_output_file):
            try:
                os.remove(self.temp_output_file)
            except Exception as e:
                print(f"⚠ Fail to remove: {e}")

    def finalize(self):
        if self.cancel_mode or self.stop_requested:
            if os.path.exists(self.temp_output_file):
                try:
                    os.remove(self.temp_output_file)
                except Exception as e:
                    print(f"⚠️ Fail to remove: {e}")
            self.signals.finished.emit("Canceled")
        else:
            self.signals.finished.emit(self.temp_output_file)

    def run(self):
        try:
            pythoncom.CoInitialize()
            self.signals.progress.emit(2)
            df = pd.read_excel(self.file_path) if self.file_path.endswith(".xlsx") else pd.read_csv(self.file_path)
            self.signals.progress.emit(5)

            total_rows = len(df)
            for i, row in df.iterrows():
                if self.stop_requested:
                    break
                self.signals.progress.emit(5 + int((i + 1) / total_rows * 45))

            today = datetime.today()
            billing_month = today.replace(day=1) - timedelta(days=1)
            year = billing_month.year
            month = billing_month.month
            month_name = f"{year}년 {month}월"

            if self.customer == "CustomerA":
                if self.stop_requested:
                    self.finalize()
                    return
                
                df_filtered = df[df["CustomerName"] == "CustomerA"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]
                df_filtered["BillingPreTaxTotal"] = df_filtered["BillingPreTaxTotal"] * 1.15

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerA Pivot"
                    )

                    for field in ["CustomerName", "SubscriptionId", "MeterSubCategory", "MeterName"]:
                        pivot_table.PivotFields(field).Orientation = 1

                    pivot_table.AddDataField(
                        pivot_table.PivotFields("BillingPreTaxTotal"),
                        "합계 BillingPreTaxTotal",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    del wb, ws, pivot_ws
                    pythoncom.CoUninitialize()
                    gc.collect()
                
                self.finalize()
                return
                    
            elif self.customer == "CustomerB":
                if self.stop_requested:
                    self.finalize()
                    return
                
                df_filtered = df[df["CustomerName"] == "CustomerB"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerB"
                    )

                    for field in ["CustomerName", "MeterCategory", "MeterSubCategory", "MeterName"]:
                        pivot_table.PivotFields(field).Orientation = 1

                    pivot_table.AddDataField(
                        pivot_table.PivotFields("BillingPreTaxTotal"),
                        "합계 BillingPreTaxTotal",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    del wb, ws, pivot_ws
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerC":
                if self.stop_requested:
                    self.finalize()
                    return
                
                df_filtered = df[df["CustomerName"] == "CustomerC"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]
                file_name = f"CustomerC {billing_month.year % 100}년 {billing_month.month:02d}월 Azure 사용량.xlsx"

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerCPivot"
                    )

                    for field in ["CustomerName", "SubscriptionId", "MeterCategory", "MeterSubCategory", "MeterName"]:
                        pivot_table.PivotFields(field).Orientation = 1

                    pivot_table.AddDataField(
                        pivot_table.PivotFields("BillingPreTaxTotal"),
                        "합계 BillingPreTaxTotal",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    del wb, ws, pivot_ws
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerD":
                if self.stop_requested:
                    self.finalize()
                    return

                df_filtered = df[df["CustomerName"] == "CustomerD"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerD_Pivot"
                    )

                    # 필터 (EntitlementDescription)
                    pivot_table.PivotFields("EntitlementDescription").Orientation = 3  # Filter

                    # 행
                    for field in ["MeterCategory", "MeterName"]:
                        pivot_table.PivotFields(field).Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("BillingPreTaxTotal"),
                        "합계 BillingPreTaxTotal",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    del wb, ws, pivot_ws
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerE":
                if self.stop_requested:
                    self.finalize()
                    return

                df_filtered = df[df["CustomerName"] == "CustomerE"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    last_row = ws.UsedRange.Rows.Count
                    last_col = ws.UsedRange.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A1"),
                        TableName="CustomerEPivot"
                    )

                    # 필터
                    pivot_table.PivotFields("CustomerName").Orientation = 3  # Filter
                    pivot_table.PivotFields("CustomerName").CurrentPage = "CustomerE"

                    # 행
                    for field in ["MeterCategory", "MeterName"]:
                        pivot_table.PivotFields(field).Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("BillingPreTaxTotal"),
                        "합계 BillingPreTaxTotal",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    # 피벗 차트 생성
                    chart = pivot_ws.Shapes.AddChart2(
                        201,  # Clustered Column = xlColumnClustered
                        51,    # xlChartInPlace
                        250, 50, 500, 200  # (left, top, width, height)
                    ).Chart
                    chart.SetSourceData(pivot_table.TableRange1)
                    chart.ChartTitle.Text = "CustomerE"
                    chart.HasLegend = True
                    chart.Parent.Top = pivot_ws.Range("D3").Top
                    chart.Parent.Left = pivot_ws.Range("D3").Left
                    chart.Axes(2).MinimumScaleIsAuto = True
                    chart.Axes(2).MaximumScaleIsAuto = True
                    chart.Axes(2).MajorUnit = 1000

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerF":
                if self.stop_requested:
                    self.finalize()
                    return

                df_filtered = df.copy()

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A1"),
                        TableName="CustomerFPivot"
                    )

                    # 필터
                    pivot_table.PivotFields("청구계정이름 (BillingAccountName)").Orientation = 3
                    pivot_table.PivotFields("청구계정이름 (BillingAccountName)").CurrentPage = "CustomerF"

                    # 행
                    pivot_table.PivotFields("미터범주 (MeterCategory)").Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157
                    )
                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    # 차트
                    chart = pivot_ws.Shapes.AddChart2(
                        201,  # Clustered Column = xlColumnClustered
                        51,    # xlChartInPlace
                        250, 50, 450, 300  # (left, top, width, height)
                    ).Chart
                    chart.SetSourceData(pivot_table.TableRange1)
                    chart.ChartTitle.Text = "CustomerF"
                    chart.HasLegend = True
                    chart.Parent.Top = pivot_ws.Range("D3").Top
                    chart.Parent.Left = pivot_ws.Range("D3").Left
                    chart.Axes(2).MinimumScaleIsAuto = True
                    chart.Axes(2).MaximumScaleIsAuto = True
                    chart.Axes(2).MajorUnit = 100000

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return

            elif self.customer == "CustomerG":
                if self.stop_requested:
                    self.finalize()
                    return

                sheet_name = f"{month_name} Azure 사용량"
                df.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A1"),
                        TableName="CustomerGPivot"
                    )

                    # 필터
                    pivot_table.PivotFields("구독이름 (SubscriptionName)").Orientation = 3
                    pivot_table.PivotFields("구독이름 (SubscriptionName)").CurrentPage = "CustomerG"

                    pivot_table.PivotFields("청구프로필이름 (BillingProfileName)").Orientation = 3
                    pivot_table.PivotFields("청구프로필이름 (BillingProfileName)").CurrentPage = "CustomerG"

                    pivot_table.PivotFields("청구프로필Id (BillingProfileId)").Orientation = 3
                    pivot_table.PivotFields("청구프로필Id (BillingProfileId)").CurrentPage = 58075352

                    # 열
                    pivot_table.PivotFields("날짜 (Date)").Orientation = 2  # Column

                    # 행
                    pivot_table.PivotFields("제품 (Product)").Orientation = 1
                    pivot_table.PivotFields("수량 (Quantity)").Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157
                    )

                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerH":
                if self.stop_requested:
                    self.finalize()
                    return

                df_filtered = df[df["계정소유자Id (AccountOwnerId)"] == "CustomerH"].copy()

                sheet_name = f"{month_name} Azure 사용량"
                df_filtered.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A1"),
                        TableName="CustomerHPivot"
                    )

                    # 필터
                    pivot_table.PivotFields("구독이름 (SubscriptionName)").Orientation = 3
                    pivot_table.PivotFields("구독이름 (SubscriptionName)").CurrentPage = "CustomerH"

                    # 행
                    pivot_table.PivotFields("미터범주 (MeterCategory)").Orientation = 1
                    pivot_table.PivotFields("요금제이름 (MeterName)").Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157
                    )

                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return

            elif self.customer == "CustomerI":
                if self.stop_requested:
                    self.finalize()
                    return

                sheet_name = f"{month_name} Azure 사용량"
                df.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerIPivot"
                    )

                    # 행 필드
                    for field in [
                        "청구계정이름 (BillingAccountName)",
                        "구독이름 (SubscriptionName)",
                        "미터범주 (MeterCategory)",
                        "미터하위범주 (MeterSubCategory)",
                        "요금제이름 (MeterName)"
                    ]:
                        pivot_table.PivotFields(field).Orientation = 1

                    # 값
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157
                    )

                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return

            elif self.customer == "CustomerJ":
                if self.stop_requested:
                    self.finalize()
                    return

                df_marked = df.copy()
                for col in ["유효가격 (EffectivePrice)", "비용 (Cost)", "단가 (UnitPrice)"]:
                    if col in df_marked.columns:
                        df_marked[col] = df_marked[col].apply(
                            lambda x: x * 1.07 if isinstance(x, (int, float)) else x
                        )

                sheet_name = f"{month_name} Azure 사용량"
                df_marked.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerJPivot"
                    )

                    for field in [
                        "청구계정이름 (BillingAccountName)",
                        "구독이름 (SubscriptionName)",
                        "미터범주 (MeterCategory)",
                        "미터하위범주 (MeterSubCategory)",
                        "요금제이름 (MeterName)"
                    ]:
                        pivot_table.PivotFields(field).Orientation = 1

                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157
                    )

                    pivot_table.TableStyle2 = "PivotStyleLight20"
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return
            
            elif self.customer == "CustomerK":
                if self.stop_requested:
                    self.finalize()
                    return

                sheet_name = f"{month_name} Azure 사용량"
                df.to_excel(self.temp_output_file, index=False, sheet_name=sheet_name)

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(self.temp_output_file)

                try:
                    ws = wb.Sheets(sheet_name)
                    used_range = ws.UsedRange
                    last_row = used_range.Rows.Count
                    last_col = used_range.Columns.Count

                    header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                    header_range.Borders.LineStyle = -4142
                    header_range.Font.Bold = False

                    pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    pivot_ws.Name = f"{month_name} Pivot"

                    pivot_cache = wb.PivotCaches().Create(
                        SourceType=1,
                        SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                    )
                    pivot_table = pivot_cache.CreatePivotTable(
                        TableDestination=pivot_ws.Range("A3"),
                        TableName="CustomerKPivot"
                    )

                    # 행 필드
                    for field in [
                        "청구계정이름 (BillingAccountName)",
                        "구독이름 (SubscriptionName)",
                        "미터범주 (MeterCategory)",
                        "미터하위범주 (MeterSubCategory)",
                        "요금제이름 (MeterName)"
                    ]:
                        pivot_table.PivotFields(field).Orientation = 1

                    # 값 필드
                    pivot_table.AddDataField(
                        pivot_table.PivotFields("비용 (Cost)"),
                        "합계 비용 (Cost)",
                        -4157  # xlSum
                    )

                    # 스타일
                    pivot_table.TableStyle2 = "PivotStyleLight20"

                    # 값 포맷 (₩)
                    for col in pivot_table.DataBodyRange.Columns:
                        col.NumberFormat = "₩#,##0"

                    wb.Save()
                finally:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    gc.collect()

                self.finalize()
                return

            elif self.customer == "CustomerL":
                if self.stop_requested:
                    self.finalize()
                    return
                
                df_filtered = df[df["CustomerName"] == "CustomerL"].copy()
                df_filtered = df_filtered.loc[:, "PartnerId":"BenefitType"]

                insert_index = df_filtered.columns.get_loc("BillingPreTaxTotal")
                df_filtered.insert(insert_index, "BillingTotal", df_filtered["BillingPreTaxTotal"] * 1.1)
                df_filtered.loc[0, "BillingTotal"] = "BillingTotal"  # 첫 행은 문자열로 고정

                df_filtered["ResourceGroup"] = df_filtered["ResourceGroup"].str.lower()

                merge_temp_path = os.path.join(tempfile.gettempdir(), f"MergeCustomerL_{month}.xlsx")
                df_filtered.to_excel(merge_temp_path, index=False)

                groups = {
                    "CustomerL": "CustomerL {month}월 비용보고서(CustomerL).xlsx",
                    "CustomerL-1": "CustomerL {month}월 비용보고서(CustomerL).xlsx",
                    "CustomerL-2": "CustomerL {month}월 비용보고서(CustomerL).xlsx"
                }

                save_dir = QFileDialog.getExistingDirectory(None, "저장 폴더 선택")
                if not save_dir:
                    raise Exception("저장 폴더가 선택되지 않았습니다.")

                for rg, file_pattern in groups.items():
                    sub_df = df_filtered[df_filtered["ResourceGroup"] == rg]
                    if sub_df.empty:
                        continue

                    if rg == "CustomerL-2" and sub_df["PricingPreTaxTotal"].sum() <= 0:
                        continue

                    file_name = file_pattern.format(month=month)
                    save_path = os.path.join(save_dir, file_name)
                    sub_df.to_excel(save_path, index=False, sheet_name=f"{month_name} Azure 사용량")

                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    wb = excel.Workbooks.Open(save_path)

                    try:
                        ws = wb.Sheets(f"{month_name} Azure 사용량")
                        last_row = ws.UsedRange.Rows.Count
                        last_col = ws.UsedRange.Columns.Count

                        header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                        header_range.Borders.LineStyle = -4142
                        header_range.Font.Bold = False

                        pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                        pivot_ws.Name = f"{month_name} Pivot"

                        pivot_cache = wb.PivotCaches().Create(
                            SourceType=1,
                            SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                        )
                        pivot_table = pivot_cache.CreatePivotTable(
                            TableDestination=pivot_ws.Range("A3"),
                            TableName=f"{rg}_Pivot"
                        )

                        for field in ["CustomerName", "ResourceGroup", "MeterCategory", "MeterSubCategory", "MeterName"]:
                            pivot_table.PivotFields(field).Orientation = 1

                        pivot_table.AddDataField(
                            pivot_table.PivotFields("BillingTotal"),
                            "합계 BillingTotal",
                            -4157
                        )
                        pivot_table.TableStyle2 = "PivotStyleLight20"

                        for col in pivot_table.DataBodyRange.Columns:
                            col.NumberFormat = "₩#,##0"

                        wb.Save()
                    finally:
                        wb.Close(SaveChanges=True)
                        excel.Quit()
                        pythoncom.CoUninitialize()
                        gc.collect()

                if os.path.exists(merge_temp_path):
                    os.remove(merge_temp_path)

                self.signals.finished.emit("CustomerL Completed")
                return

        except Exception as e:
            self.signals.error.emit(str(e))
            self.signals.finished.emit("Canceled. Please choose customer again.")

    def stop(self):
        self.stop_requested = True
        self.cancel_mode = True

class BillingMasterApp(QWidget):
    def __init__(self):
        super().__init__()
        icon_path = r"C:\Projects\BillingMaster\Logo.ico"
        self.icon  = QIcon(icon_path)
        self.setWindowIcon(self.icon)
        self.setWindowTitle("Billing Master")
        self.setFixedSize(350, 500)
        self.move_to_primary_screen()

        self.layout = QVBoxLayout()

        self.title = QLabel("Billing Master", self)
        self.title.setAlignment(Qt.AlignCenter)
        self.title.setStyleSheet("font-size: 45px; font-family: 'Gabriola'; font-weight: bold;")
        self.layout.addWidget(self.title)

        outlook_btn = QPushButton("Send Email through Outlook")
        outlook_btn.setFixedHeight(30)
        outlook_btn.clicked.connect(self.show_outlook_client_selector)
        self.layout.addWidget(outlook_btn)

        self.customers = ["CustomerA", "CustomerB", "CustomerC", "CustomerD", "CustomerE", "CustomerF", "CustomerG", "CustomerH", "CustomerI", "CustomerJ", "CustomerK", "CustomerL", "CustomerM", "CustomerN"]
        grid_layout = QGridLayout()
        grid_layout.setHorizontalSpacing(20)
        grid_layout.setVerticalSpacing(10)

        for i, customer in enumerate(self.customers):
            btn = QPushButton(customer)
            btn.setFixedSize(120, 40)
            btn.clicked.connect(lambda _, c=customer: self.open_file_dialog(c))
            row = i // 2
            col = i % 2
            grid_layout.addWidget(btn, row, col)

        self.layout.addLayout(grid_layout)

        self.notice_label = QLabel("", self)
        self.notice_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #0044aa; padding: 4px 0;")
        self.layout.addWidget(self.notice_label)

        status_row = QHBoxLayout()
        self.status_label = QLabel("", self)
        self.status_label.setStyleSheet("font-size: 11px; margin: 0px; padding: 0px;")
        self.cancel_button = QPushButton("취소", self)
        self.cancel_button.setVisible(False)
        self.cancel_button.setFixedSize(50, 25)
        self.cancel_button.clicked.connect(self.cancel_conversion)
        status_row.addWidget(self.status_label)
        status_row.addWidget(self.cancel_button)
        status_row.setSpacing(10)

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)

        status_layout = QVBoxLayout()
        status_layout.setContentsMargins(0, 0, 0, 0)
        status_layout.setSpacing(3)
        status_layout.addLayout(status_row)
        status_layout.addWidget(self.progress_bar)

        status_widget = QWidget()
        status_widget.setLayout(status_layout)
        self.layout.addWidget(status_widget)
        self.setLayout(self.layout)

    def move_to_primary_screen(self):
        screen = QApplication.primaryScreen()
        geometry = screen.availableGeometry()
        self.move(geometry.left(), geometry.top())

    def show_outlook_client_selector(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Which customer do you want to send?")
        dialog.resize(300, 200)

        layout = QVBoxLayout()
        customer_list = [
            "CustomerA", "CustomerB", "CustomerC", "CustomerD", "CustomerE",
            "CustomerF", "CustomerG", "CustomerH",
            "CustomerI", "스CustomerJ", "CustomerK", "CustomerL"
        ]

        for customer in customer_list:
            btn = QPushButton(customer)
            btn.clicked.connect(lambda _, c=customer: (dialog.accept(), self.send_outlook_email(c)))
            layout.addWidget(btn)

        dialog.setLayout(layout)
        dialog.exec()

    def send_outlook_email(self, customer):
        today = datetime.today()
        billing_month = today.replace(day=1) - timedelta(days=1)
        send_year = billing_month.year
        send_month = billing_month.month

        if customer == "CustomerA":
            filename = f"CustomerA {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            file_path, _ = QFileDialog.getOpenFileName(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not file_path:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerA] {send_year % 100}년 {send_month:02d}월 Azure 사용량 송부 건"

                mail.Attachments.Add(file_path)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    Mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerB":
            filename = f"CustomerB {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerB] {send_month:02d}월 비용보고서&점검 보고서 전달드립니다"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerC/D":
            filename = f"CustomerC/D {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerC/D] {send_year % 100}.{send_month:02d}월 빌링 안내"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerE":
            filename = f"CustomerE {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerE] {send_year % 100}년 {send_month:02d}월 사용량 보고서"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerF":
            filename = f"CustomerF {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerF] {send_year % 100}년 {send_month:02d}월 Azure 사용량 송부의 건"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerG":
            filename = f"CustomerG {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerG] Azure {send_year % 100}년 {send_month:02d}월 사용량 파일 전달"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerH":
            filename = f"CustomerH {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerH] Azure {send_year % 100}년 {send_month:02d}월 한달 사용비용"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerI":
            filename = f"CustomerI {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerI] {send_year % 100}.{send_month:02d} Microsoft Azure 사용량"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerJ":
            filename = f"CustomerJ {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[Azure 청구 금액] CustomerJ {send_year % 100}년 {send_month:02d}월"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerK":
            filename = f"CustomerK {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerK] {send_year % 100}년 {send_month:02d}월 사용 비용"

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerL":
            filename = f"CustomerL {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerL] {send_month:02d}월 비용보고서 전달드립니다."

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

        elif customer == "CustomerM":
            filename = f"CustomerM {send_year}년 {send_month:02d}월 Azure 사용량.xlsx"
            files, _ = QFileDialog.getOpenFileNames(self, "첨부할 파일을 선택하세요", filename, "Excel Files (*.xlsx)")
            if not files:
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "abcd@abcd.com"
                mail.CC = "abcd@abcd.com"
                mail.Subject = f"[CustomerM] {send_month:02d}월 비용보고서 전달드립니다."

                for f in files:
                    mail.Attachments.Add(f)
                mail.Display(False)

                for _ in range(50):
                    pythoncom.PumpWaitingMessages()   # GUI 메세지 처리
                    if mail.HTMLBody and "</html>" in mail.HTMLBody.lower():
                        break
                    time.sleep(0.1)

                signature = mail.HTMLBody
                
                custom_body = f"""
                <div style="font-family:'맑은 고딕'; font-size:10pt;">
                    mail content
                </div>
                """

                mail.HTMLBody = custom_body + signature

                mail.Display(True)

            except Exception as e:
                QMessageBox.critical(self, "오류", f"Outlook 실행 실패: {str(e)}")

    def open_file_dialog(self, customer):
        self.customer = customer
        self.status_label.setText("")
        self.notice_label.setText("")

        if customer == "CustomerJ":
            self.notice_label.setText("📌 <b>PEC 자료를 먼저 Upload 하세요.</b>")
            self.notice_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #0044aa;")
            file_path, _ = QFileDialog.getOpenFileName(self, "PEC 파일 업로드", "", "Excel or CSV Files (*.xlsx *.csv)")
            if file_path:
                self.CustomerJ(file_path)
            return
        
        elif customer == "CustomerK":
            dlg = QDialog(self)
            dlg.setWindowTitle("CustomerK 선택")
            dlg.resize(self.width() * 0.5, self.height() * 0.3)

            btn_ea  = QPushButton("EA 빌링",  dlg)
            btn_csp = QPushButton("CSP 빌링", dlg)
            for b in (btn_ea, btn_csp):
                b.setMinimumHeight(60)
                b.setStyleSheet("font-size: 13px;")

            lay = QHBoxLayout()
            lay.addWidget(btn_ea)
            lay.addWidget(btn_csp)
            dlg.setLayout(lay)

            choice = {}
            btn_ea.clicked.connect(lambda: (choice.setdefault("v", "EA"), dlg.accept()))
            btn_csp.clicked.connect(lambda: (choice.setdefault("v", "CSP"), dlg.accept()))

            if dlg.exec() != QDialog.Accepted:
                return

            if choice.get("v") == "EA":
                self.CustomerK()   # ← 방금 만든 메서드 호출
            else:
                self.CustomerK()   # (기존 CSP 로직이 있다면 호출)
            return
        
        self.status_label.setText("")
        file_path, _ = QFileDialog.getOpenFileName(self, "파일 업로드", "", "Excel or CSV Files (*.xlsx *.csv)")
        if file_path:
            if not (file_path.endswith(".csv") or file_path.endswith(".xlsx")):
                QMessageBox.warning(self, "오류", "파일 형식이 맞지 않습니다.")
                return
            self.file_path = file_path
            self.start_conversion()

    def CustomerK(self, file_path):
        try:
            self.status_label.setText("PEC 변환 중입니다...")
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(5)
            self.cancel_button.setVisible(True)

            df = pd.read_excel(file_path) if file_path.endswith(".xlsx") else pd.read_csv(file_path)
            df_filtered = df[df["CustomerName"] == "CustomerB"].copy()
            if df_filtered.empty:
                raise Exception("CustomerB 데이터가 없습니다.")

            self.progress_bar.setValue(10)

            today = datetime.today()
            billing_month = today.replace(day=1) - timedelta(days=1)
            year_month = billing_month.strftime("%Y%m")
            self.cw_save_dir = QFileDialog.getExistingDirectory(self, "저장 위치 선택")
            if not self.cw_save_dir:
                raise Exception("저장 위치가 선택되지 않았습니다.")

            temp_file = os.path.join(tempfile.gettempdir(), f"cw_pec_{year_month}.xlsx")
            self.temp_output_file = temp_file
            df_filtered.to_excel(temp_file, index=False, sheet_name="Data")

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(temp_file)

            try:
                ws = wb.Sheets("Data")
                last_row = ws.UsedRange.Rows.Count
                last_col = ws.UsedRange.Columns.Count

                header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                header_range.Borders.LineStyle = -4142
                header_range.Font.Bold = False

                pivot_ws = wb.Sheets.Add(After=ws)
                pivot_ws.Name = "Summary"

                pivot_cache = wb.PivotCaches().Create(
                    SourceType=1,
                    SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                )
                pivot_table = pivot_cache.CreatePivotTable(
                    TableDestination=pivot_ws.Range("A3"),
                    TableName="CW_PEC_Pivot"
                )

                pivot_table.PivotFields("EntitlementDescription").Orientation = 3  # Filter
                pivot_table.PivotFields("EntitlementDescription").CurrentPage = "CustomerB"

                for field in ["MeterCategory", "MeterName"]:
                    pivot_table.PivotFields(field).Orientation = 1

                pivot_table.AddDataField(
                    pivot_table.PivotFields("BillingPreTaxTotal"),
                    "합계 BillingPreTaxTotal",
                    -4157
                )
                pivot_table.TableStyle2 = "PivotStyleLight20"
                for col in pivot_table.DataBodyRange.Columns:
                    col.NumberFormat = "₩#,##0"

                wb.Save()
            finally:
                wb.Close(SaveChanges=True)
                excel.Quit()
                pythoncom.CoUninitialize()
                gc.collect()

            save_path = os.path.join(self.cw_save_dir, f"{year_month}_CustomerB.xlsx")
            os.replace(temp_file, save_path)
            self.status_label.setText("PEC 저장 완료. CostManagement 파일 업로드 하세요.")

            self.progress_bar.setVisible(False)
            self.cancel_button.setVisible(False)
            self.temp_output_file = None
            self.notice_label.setText("📌 <b>CostManagement 파일을 Upload 하세요.</b>")
            self.notice_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #0044aa;")

            self.CustomerK()

        except Exception as e:
            if self.temp_output_file and os.path.exists(self.temp_output_file):
                os.remove(self.temp_output_file)
            self.status_label.setText("취소가 완료되었습니다.")
            self.progress_bar.setVisible(False)
            self.cancel_button.setVisible(False)
            self.notice_label.setText("")

    def CustomerK(self):
        try:
            self.status_label.setText("CostManagement 변환 중입니다...")
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(5)
            self.cancel_button.setVisible(True)

            file_path, _ = QFileDialog.getOpenFileName(self, "CostManagement 파일 업로드", "", "Excel or CSV Files (*.xlsx *.csv)")
            if not file_path:
                raise Exception("파일이 선택되지 않았습니다.")

            df = pd.read_excel(file_path, sheet_name="Data") if file_path.endswith(".xlsx") else pd.read_csv(file_path)
            self.progress_bar.setValue(10)

            temp_file = os.path.join(tempfile.gettempdir(), f"cw_cost_temp_{uuid.uuid4().hex}.xlsx")
            self.temp_output_file = temp_file
            df.to_excel(temp_file, index=False, sheet_name="Data")

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(temp_file)

            try:
                ws = wb.Sheets("Data")
                last_row = ws.UsedRange.Rows.Count
                last_col = ws.UsedRange.Columns.Count

                header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
                header_range.Borders.LineStyle = -4142
                header_range.Font.Bold = False

                pivot_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                pivot_ws.Name = "Billing"

                pivot_cache = wb.PivotCaches().Create(
                    SourceType=1,
                    SourceData=ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
                )
                pivot_table = pivot_cache.CreatePivotTable(
                    TableDestination=pivot_ws.Range("A3"),
                    TableName="CW_Cost_Pivot"
                )

                for field in ["ServiceName", "Meter"]:
                    pivot_table.PivotFields(field).Orientation = 1

                pivot_table.AddDataField(
                    pivot_table.PivotFields("Cost"),
                    "합계 Cost",
                    -4157
                )
                pivot_table.TableStyle2 = "PivotStyleLight20"
                for col in pivot_table.DataBodyRange.Columns:
                    col.NumberFormat = "₩#,##0"

                wb.Save()
            finally:
                wb.Close(SaveChanges=True)
                excel.Quit()
                pythoncom.CoUninitialize()
                gc.collect()

            save_path = os.path.join(self.cw_save_dir, os.path.basename(file_path))
            os.replace(temp_file, save_path)

            self.status_label.setText("작업이 완료되었습니다.")
            self.progress_bar.setVisible(False)
            self.cancel_button.setVisible(False)
            self.temp_output_file = None
            self.notice_label.setText("")

        except Exception as e:
            if self.temp_output_file and os.path.exists(self.temp_output_file):
                os.remove(self.temp_output_file)
            self.status_label.setText("취소가 완료되었습니다.")
            self.progress_bar.setVisible(False)
            self.cancel_button.setVisible(False)
            self.notice_label.setText("")

    def CustomerK(self):
        try:
            src, _ = QFileDialog.getOpenFileName(
            self,
            "EA Billing 파일 업로드",
            "",
            "Excel and CSV Files (*.xlsx *.xls *.csv);;All Files (*)"
            )
            if not src:
                return
            
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(15)
            self.status_label.setText("🔍 데이터 로드 중…")
            
            ext = os.path.splitext(src)[1].lower()
            if ext in [".xlsx", ".xls"]:
                df = pd.read_excel(src, engine="openpyxl")
            else:           # csv
                try:
                    df = pd.read_csv(src, encoding="utf-8")
                except UnicodeDecodeError:
                    df = pd.read_csv(src, encoding="cp949")

            df = df[df["계정이름 (AccountName)"] == "CustomerK"]
            df["구독이름 (SubscriptionName)"] = (
            df["구독이름 (SubscriptionName)"].str.strip().str.lower()
            )
            if df.empty:
                QMessageBox.information(self, "데이터 없음", "EA 대상 계정 데이터가 없습니다.")
                self.progress_bar.setVisible(False)
                return
            
            dst_dir = QFileDialog.getExistingDirectory(self, "저장 폴더 선택")
            if not dst_dir:
                self.progress_bar.setVisible(False)
                return

            bill_month = datetime.today().replace(day=1) - timedelta(days=1)
            final_path = os.path.join(
                dst_dir, f"CustomerB {bill_month.year}{bill_month.month:02d} 비용.xlsx")
            yymm = f"{bill_month.year%100:02d}{bill_month.month:02d}"
            main_sheet  = f"CustomerB{yymm}_비용데이터"

            self.progress_bar.setValue(30)
            self.status_label.setText("📑 Excel 세션 생성…")
            
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            tmp_xlsx = os.path.join(tempfile.gettempdir(),
                        f"CustomerB{os.getpid()}.xlsx")
            wb_tmp = openpyxl.Workbook()
            ws_tmp = wb_tmp.active
            ws_tmp.title = main_sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_tmp.append(r)
            wb_tmp.save(tmp_xlsx)
            wb_tmp.close()

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(tmp_xlsx)
            ws_main = wb.Sheets(main_sheet)

            piv_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
            piv_ws.Name = "CustomerB"
            piv_ws.Range("A2").Value = "1. CustomerB"
            a2 = piv_ws.Range("A2")
            a2.Font.Name, a2.Font.Size, a2.Font.Bold = "맑은 고딕", 14, True

            cache = wb.PivotCaches().Create(SourceType=1, SourceData=ws_main.UsedRange)
            pvt   = cache.CreatePivotTable(TableDestination=piv_ws.Range("A3"), TableName="EA_Summary")
            pvt.PivotFields("구독이름 (SubscriptionName)").Orientation = 1
            pvt.AddDataField(pvt.PivotFields("비용 (Cost)"), "합계 Cost", -4157)
            pvt.TableStyle2 = "PivotStyleLight20"
            for col in pvt.DataBodyRange.Columns:
                col.NumberFormat = "₩#,##0"

            sub_col = ws_main.Range("1:1").Find("구독이름 (SubscriptionName)").Column
            def split_and_pivot(sub, cell):
                sub_df = df[df["구독이름 (SubscriptionName)"]
                  .str.strip().str.lower().eq(sub)]
                if sub_df.empty:
                    return False
                
                new_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                new_ws.Name = sub
                for r_idx, r in enumerate(dataframe_to_rows(sub_df, index=False,
                                                header=True), start=1):
                    for c_idx, v in enumerate(r, start=1):
                        new_ws.Cells(r_idx, c_idx).Value = v

                cache = wb.PivotCaches().Create(SourceType=1,
                                    SourceData=new_ws.UsedRange)
                pv = cache.CreatePivotTable(TableDestination=piv_ws.Range(cell),
                                TableName=f"{sub}_Pivot")
                pv.PivotFields("리소스그룹 (ResourceGroup)").Orientation = 1
                pv.AddDataField(pv.PivotFields("비용 (Cost)"), "합계 Cost", -4157)
                pv.TableStyle2 = "PivotStyleLight20"
                for col in pv.DataBodyRange.Columns:
                    col.NumberFormat = "₩#,##0"
                return True

            for sub, where in [
                    ("CustomerB", "A15"),
                    ("CustomerB-1", "E15"),
                    ("CustomerB-2", "I15")]:
                split_and_pivot(sub, where)

            piv_ws.Range("A14").Value = "2. CustomerB"
            order = [
                "CustomerB",                 # 요약 시트
                main_sheet,                   
                "CustomerB",
                "CustomerB-1",
                "CustomerB-2",
            ]
            last_pos = 1
            for name in order[::-1]:          # 뒤에서부터 앞으로 이동하면 index 어긋남 없음
                try:
                    wb.Worksheets(name).Move(Before=wb.Worksheets(1))
                except Exception:
                    pass 
            a14 = piv_ws.Range("A14")
            a14.Font.Name, a14.Font.Size, a14.Font.Bold = "맑은 고딕", 14, True

            self.progress_bar.setValue(90)
            self.status_label.setText("💾 최종 저장 중…")

            if os.path.exists(final_path):
                os.remove(final_path)

            tmp_save = os.path.join(tempfile.gettempdir(),
                                    f"CustomerK_{uuid.uuid4().hex}.xlsx")
            wb.SaveAs(tmp_save)
            wb.Close(SaveChanges=False)
            excel.Quit()
            pythoncom.CoUninitialize()

            self.progress_bar.setValue(95)
            self.status_label.setText("🗜️  파일 이동 중…")
            dst_dir   = urllib.parse.unquote(dst_dir)
            final_path = pathlib.Path(dst_dir) / \
                         f"CustomerB{bill_month:%Y%m}비용.xlsx"
            if final_path.exists():
                final_path.unlink()

            shutil.move(tmp_save, final_path)
            if os.path.exists(tmp_save):
                os.remove(tmp_save)
            if os.path.exists(tmp_xlsx):
                os.remove(tmp_xlsx)

            self.progress_bar.setValue(100)
            self.status_label.setText("✅ 완료!")
            QMessageBox.information(self, "완료",
                                     f"작업이 완료되었습니다.")
            
        except Exception as err:
            QMessageBox.critical(self, "EA Billing 오류", str(err))
        finally:
            self.progress_bar.setVisible(False)

    def CustomerK(self):
        try:
            paths, _ = QFileDialog.getOpenFileNames(
                self, "CSP Billing 원본 파일(들) 선택", "",
                "Excel / CSV Files (*.xlsx *.xls *.csv);;All Files (*)")
            if len(paths) < 2:
                QMessageBox.warning(self, "선택 오류",
                                    "두 개 이상의 CSP 원본 파일을 선택해 주세요.")
                return
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(10)
            self.status_label.setText("📑 원본 병합 중…")

            tmp_merge = os.path.join(tempfile.gettempdir(),"merge_result_csp.xlsx")
            wb_tmp = openpyxl.Workbook(); wb_tmp.remove(wb_tmp.active)

            for p in paths:
                ext = os.path.splitext(p)[1].lower()
                if ext in (".xlsx", ".xls"):
                    df_src = pd.read_excel(p, sheet_name="Data", engine="openpyxl")
                else:
                    try:
                        df_src = pd.read_csv(p, encoding="utf-8")
                    except UnicodeDecodeError:
                        df_src = pd.read_csv(p, encoding="cp949")
                sh = wb_tmp.create_sheet()
                for r in dataframe_to_rows(df_src, index=False, header=True):
                    sh.append(r)
            wb_tmp.save(tmp_merge); wb_tmp.close()

            self.progress_bar.setValue(25)

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application"); excel.Visible=False
            wb = excel.Workbooks.Open(tmp_merge)

            # ────────────────────────────── 시트명 정리
            for sht in wb.Sheets:
                g2 = sht.Range("G2").Value
                if g2 == "CustomerB":
                    sht.Name = "CustomerB"
                elif g2 == "CustomerB":
                    sht.Name = "CustomerB"

            ws_aoai   = wb.Sheets("CustomerB")
            ws_entraid = wb.Sheets("CustomerB")

            self.progress_bar.setValue(35); self.status_label.setText("📄 데이터 합치기…")
            y, m = (datetime.today().replace(day=1)-timedelta(days=1)).strftime("%Y"), \
                   (datetime.today().replace(day=1)-timedelta(days=1)).strftime("%m")
            data_sht_name = f"CustomerB{y}{m}_비용데이터"
            ws_data = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count)); ws_data.Name = data_sht_name

            def copy_used(src_ws, dst_ws, dst_row, skip_header=False):
                used = src_ws.UsedRange
                r_cnt = used.Rows.Count
                c_cnt = used.Columns.Count
                start = 2 if skip_header else 1
                rng = src_ws.Range(src_ws.Cells(start,1),
                                src_ws.Cells(r_cnt, c_cnt))
                rng.Copy(dst_ws.Cells(dst_row, 1))
                return r_cnt - (1 if skip_header else 0)

            rows1 = copy_used(ws_aoai, ws_data, 1)
            copy_used(ws_entraid, ws_data, rows1+1, skip_header=True)

            # ────────────────────────────── Pivot 시트
            self.progress_bar.setValue(50); self.status_label.setText("📊 PivotSheet 생성…")
            piv_sht_name = f"CustomerB{y}{m}"
            piv_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count)); piv_ws.Name = piv_sht_name
            piv_ws.Cells.Font.Name = "맑은 고딕"

            def make_pivot(src_range, dst_cell, t_name,
                        rows, data_field):
                pc = wb.PivotCaches().Create(1, src_range)
                pt = pc.CreatePivotTable(TableDestination=piv_ws.Range(dst_cell), TableName=t_name)
                for r in rows:
                    pt.PivotFields(r).Orientation = 1
                pt.AddDataField(pt.PivotFields(data_field[0]),
                                data_field[1], -4157)
                for col in pt.DataBodyRange.Columns:
                    col.NumberFormat = "₩#,##0"
                pt.TableStyle2 = "PivotStyleLight20"

            piv_ws.Range("A2").Value = "1. CustomerB"
            piv_ws.Range("A2").Font.Bold, piv_ws.Range("A2").Font.Size = True,14
            make_pivot(ws_data.UsedRange,"A3","PivotSumCost",
                       ["SubscriptionName"],("Cost","합계 Cost"))

            piv_ws.Range("A12").Value="2. CustomerB"
            piv_ws.Range("A12").Font.Bold, piv_ws.Range("A12").Font.Size = True,14

            make_pivot(ws_entraid.UsedRange,"A13","PivotEntraid",
                       ["SubscriptionName","ServiceName","Product"],("Cost","합계 Cost"))
            make_pivot(ws_aoai.UsedRange,"D13","PivotAOAI",
                       ["SubscriptionName","ServiceName","Product"],("Cost","합계 Cost"))

            # ────────────────────────────── 시트 순서
            for nm in [piv_sht_name, data_sht_name,
                       "CustomerB","CustomerB1"][::-1]:
                wb.Worksheets(nm).Move(Before=wb.Worksheets(1))

            self.progress_bar.setValue(70); self.status_label.setText("💾 임시 저장…")
            tmp_save = os.path.join(tempfile.gettempdir(), f"CustomerB_{uuid.uuid4().hex}.xlsx")
            wb.SaveAs(tmp_save); wb.Close(SaveChanges=False); excel.Quit(); pythoncom.CoUninitialize()

            # ────────────────────────────── 최종 저장
            self.progress_bar.setValue(85); self.status_label.setText("📂 저장 위치 선택…")
            target, _ = QFileDialog.getSaveFileName(self,"최종 파일 저장",
                          f"CustomerB {int(m)}월 비용.xlsx","Excel Files (*.xlsx)")
            if not target:
                os.remove(tmp_save); os.remove(tmp_merge); self.progress_bar.setVisible(False); return
            shutil.move(tmp_save, target)

            # temp 파일 정리
            for f in (tmp_save, tmp_merge):
                if os.path.exists(f): os.remove(f)

            self.progress_bar.setValue(100)
            self.status_label.setText("✅ 완료!")
            QMessageBox.information(self,"완료",f"작업이 완료되었습니다.")

        except Exception as err:
            QMessageBox.critical(self,"CSP 오류",str(err))
            for f in ("tmp_save","tmp_merge"):
                try:
                    path = locals().get(f); 
                    if path and os.path.exists(path): os.remove(path)
                except: pass
        finally:
            self.progress_bar.setVisible(False)

    def start_conversion(self):
        self.status_label.setText("변환 중입니다...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.cancel_button.setVisible(True)

        self.signals = WorkerSignals()
        self.signals.progress.connect(self.progress_bar.setValue)
        self.signals.finished.connect(self.conversion_done)
        self.signals.error.connect(self.show_error)

        self.worker = BillingWorker(self.customer, self.file_path, self.signals)
        self.worker.start()

    def conversion_done(self, temp_file):
        self.progress_bar.setVisible(False)
        self.cancel_button.setVisible(False)

        if temp_file == "Canceled":
            self.status_label.setText("취소가 완료되었습니다. 고객사를 다시 선택해주세요")
            return

        if self.customer == "CustomerB" and temp_file == "CustomerB 완료":
            self.status_label.setText("작업이 완료되었습니다.")
            return

        today = datetime.today()
        billing_month = today.replace(day=1) - timedelta(days=1)

        if self.customer == "CustomerA":
            file_name = f"iCustomerA {billing_month.year}년 {billing_month.month:02d}월 Azure 사용량.xlsx"
        elif self.customer == "CustomerB":
            file_name = f"CustomerB {billing_month.month}월 비용보고서.xlsx"
        elif self.customer == "CustomerC":
            file_name = f"CustomerC {billing_month.year % 100}년 {billing_month.month:02d}월 Azure 사용량.xlsx"
        elif self.customer == "CustomerD":
            today = datetime.today()
            file_name = f"{today.strftime('%Y%m%d')}_{billing_month.month:02d}월Billing.xlsx"
        elif self.customer == "CustomerE":
            file_name = f"CustomerF {billing_month.year}년 {billing_month.month}월 Azure 사용량.xlsx"
        elif self.customer == "CustomerG":
            file_name = f"{billing_month.year % 100}년 {billing_month.month}월 사용 금액_Stand Egg (Azure Portal).xlsx"
        elif self.customer == "CustomerH":
            today = datetime.today()
            file_name = f"CustomerI {billing_month.year}년 {billing_month.month}월_{today.strftime('%Y%m%d')}.xlsx"
        elif self.customer == "CustomerJ":
            file_name = f"CustomerK {billing_month.year}{billing_month.month:02d}비용.xlsx"
        elif self.customer == "CustomerL":
            file_name = f"CustomerM {billing_month.month}월 비용보고서.xlsx"
        elif self.customer == "CustomerN":
            file_name = f"CustomerO {billing_month.month}월 비용보고서.xlsx"
        elif self.customer == "CustomerP":
            file_name = f"CustomerQ {billing_month.month:02d}월 비용.xlsx"
        else:
            file_name = f"{self.customer} {billing_month.year}-{billing_month.month:02d}.xlsx"

        save_path, _ = QFileDialog.getSaveFileName(self, "저장 위치 선택", file_name, "Excel Files (*.xlsx)")
        if save_path:
            try:
                os.replace(temp_file, save_path)
                self.status_label.setText("작업이 완료되었습니다.")
            except Exception as e:
                self.show_error(f"파일 저장 오류: {str(e)}")

    def cancel_conversion(self):
        if hasattr(self, "worker") and self.worker.is_alive():
            self.worker.stop()
            self.status_label.setText("취소 중...")
            self.progress_bar.setVisible(True)

    def show_error(self, message):
        QMessageBox.critical(self, "에러", f"오류 발생: {message}")
        self.status_label.setText("")
        self.progress_bar.setVisible(False)
        self.cancel_button.setVisible(False)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(r"C:\Projects\BillingMaster\Logo.ico"))
    window = BillingMasterApp()
    window.show()
    sys.exit(app.exec())